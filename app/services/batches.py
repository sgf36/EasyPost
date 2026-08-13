"""Batch shipments: import/validation, bulk rate + buy, combined labels.

Recipients are imported from either a plain ``.csv`` or an Excel ``.xlsx``
workbook. The workbook template additionally carries a real dropdown on the
``predefined_package`` column — a data-validation list of carrier-qualified
package labels ("Royal Mail — LETTER"), so a code that several carriers share
still names its carrier — which a flat CSV cannot express. Either format parses
to the same :class:`BatchRow` list, and the carrier-qualified label is reduced
back to the bare code EasyPost expects at submission time.
"""

import csv
import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from app.core.client import client_manager
from app.core.customs import build_customs_info, currency_for, customs_item, is_international
from app.core.db import db_cursor
from app.services.packages import package_code_from_choice
from app.services.shipments import preferred_label_options

logger = logging.getLogger(__name__)

CSV_COLUMNS = [
    "to_name",
    "to_company",
    "to_street1",
    "to_street2",
    "to_city",
    "to_state",
    "to_zip",
    "to_country",
    "to_phone",
    "to_email",
    "length",
    "width",
    "height",
    "weight",
    # A carrier package code (e.g. "FlatRateEnvelope"). When set, the carrier's
    # own fixed dimensions apply and length/width/height are ignored — so they
    # stop being required for that row. The .xlsx template offers this as a
    # dropdown; see write_xlsx_template.
    "predefined_package",
    "reference",
    # Customs. Required only on rows that cross a border, and ignored on those
    # that do not — see _validate_row. One item per parcel, which is the shape a
    # batch almost always has; a parcel needing several declared lines is a
    # single shipment, not a spreadsheet row.
    #
    # The item's weight is taken from the parcel `weight` above rather than
    # asking twice: for a one-item parcel they are the same number, and two
    # columns that must agree is two columns that can disagree.
    "customs_description",
    "customs_quantity",
    "customs_value",
    "customs_hs_tariff",
    "customs_origin_country",
]

# Needed on any row whose destination country differs from the sender's. Sending
# an international shipment without them creates the batch and then fails every
# label at purchase, with the carrier's own wording and no mention of customs:
# "At least one item per package must be provided."
CUSTOMS_REQUIRED_COLUMNS = ("customs_description", "customs_value")

# Header-level requirement: these columns must exist. length/width/height are
# additionally required *per row* only when that row names no predefined
# package (see _validate_row) — a predefined package supplies its own.
#
# `to_state` is deliberately NOT here. It is required only for the countries
# that actually use a state/province in their addressing (see
# STATE_REQUIRED_COUNTRIES). EasyPost's own documented GB example carries no
# `state` at all, and requiring it rejected valid United Kingdom rows before
# they ever reached the API.
REQUIRED_COLUMNS = {"to_street1", "to_city", "to_zip", "to_country", "weight"}

# Countries whose addresses genuinely need a state, province or territory.
# Everywhere else it is optional, and EasyPost fills in a normalised value of
# its own where one applies.
STATE_REQUIRED_COUNTRIES = {"US", "CA", "AU", "IN", "BR", "MX", "MY", "AR"}

DIMENSION_COLUMNS = ("length", "width", "height")

_SAMPLE_ROW = {
    "to_name": "Jane Doe", "to_company": "", "to_street1": "123 Main St",
    "to_street2": "", "to_city": "Boston", "to_state": "MA", "to_zip": "02110",
    "to_country": "US", "to_phone": "5551234567", "to_email": "jane@example.com",
    "length": "10", "width": "6", "height": "4", "weight": "16",
    "predefined_package": "", "reference": "order-1001",
    # Filled in on the sample row so the columns are self-explanatory, and
    # harmless on a domestic row, where they are ignored.
    "customs_description": "Cotton t-shirt", "customs_quantity": "1",
    "customs_value": "12.50", "customs_hs_tariff": "610910",
    "customs_origin_country": "GB",
}


def write_csv_template(path: str) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(CSV_COLUMNS)
        writer.writerow([_SAMPLE_ROW[col] for col in CSV_COLUMNS])


def write_xlsx_template(path: str, package_choices: Optional[list[str]] = None) -> None:
    """Write the batch template as an Excel workbook with a package dropdown.

    ``package_choices`` populates a data-validation list on every cell of the
    ``predefined_package`` column, so the recipient sheet offers the carrier
    package codes as a real dropdown rather than free text. The choices are
    kept on a second, hidden sheet and referenced by range — an inline list
    formula is capped near 255 characters, which the full carrier set blows
    past. When no choices are available (offline, empty cache) the column is
    left as free text.
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    choices = [c for c in (package_choices or []) if c]

    wb = Workbook()
    ws = wb.active
    ws.title = "Recipients"
    ws.append(CSV_COLUMNS)
    ws.append([_SAMPLE_ROW[col] for col in CSV_COLUMNS])
    ws.freeze_panes = "A2"

    if choices:
        opts = wb.create_sheet("Packages")
        opts.append(["predefined_package"])
        for name in choices:
            opts.append([name])
        opts.sheet_state = "hidden"

        pkg_col = get_column_letter(CSV_COLUMNS.index("predefined_package") + 1)
        ref = f"Packages!$A$2:$A${len(choices) + 1}"
        dv = DataValidation(type="list", formula1=f"={ref}", allow_blank=True)
        # Apply to the data rows below the header, not the header itself.
        dv.add(f"{pkg_col}2:{pkg_col}1048576")
        ws.add_data_validation(dv)

    wb.save(path)


@dataclass
class BatchRow:
    line_number: int
    fields: dict
    errors: list[str]

    @property
    def is_valid(self) -> bool:
        return not self.errors


def _validate_row(line_number: int, fields: dict, from_country: Optional[str] = None) -> BatchRow:
    fields = {k: (str(v) if v is not None else "").strip() for k, v in fields.items()}
    errors = [col for col in REQUIRED_COLUMNS if not fields.get(col)]

    # State only where the destination country actually uses one.
    if fields.get("to_country", "").upper() in STATE_REQUIRED_COUNTRIES and not fields.get("to_state"):
        errors.append("to_state")

    # Customs, on rows that cross a border. Checked here so an incomplete row is
    # refused in the preview, before a batch exists — the alternative is what
    # this fixes: a batch created against a live account, then every label
    # failing at purchase.
    #
    # `from_country` is unknown when a file is parsed before a sender is chosen.
    # In that case the row is left alone rather than guessed at; the view
    # re-validates once the sender is known.
    if is_international(from_country, fields.get("to_country")):
        errors.extend(col for col in CUSTOMS_REQUIRED_COLUMNS if not fields.get(col))
        quantity = fields.get("customs_quantity")
        if quantity:
            try:
                if int(float(quantity)) < 1:
                    errors.append("customs_quantity must be at least 1")
            except ValueError:
                errors.append("customs_quantity is not a number")
        value = fields.get("customs_value")
        if value:
            try:
                float(value)
            except ValueError:
                errors.append("customs_value is not a number")

    # A predefined package brings its own dimensions, so length/width/height
    # are only required when no package is named. When they are supplied either
    # way, they must still be numeric.
    if not fields.get("predefined_package"):
        errors.extend(col for col in DIMENSION_COLUMNS if not fields.get(col))

    for numeric_col in (*DIMENSION_COLUMNS, "weight"):
        value = fields.get(numeric_col)
        if value:
            try:
                float(value)
            except ValueError:
                errors.append(f"{numeric_col} is not a number")

    return BatchRow(line_number=line_number, fields=fields, errors=errors)


def _check_header(fieldnames) -> None:
    missing = REQUIRED_COLUMNS - set(fieldnames or [])
    if missing:
        raise ValueError(f"File is missing required columns: {', '.join(sorted(missing))}")


def _parse_xlsx(path: str, from_country: Optional[str] = None) -> list[BatchRow]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Recipients"] if "Recipients" in wb.sheetnames else wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(it, []) or []]
    _check_header(header)

    rows: list[BatchRow] = []
    for line_number, raw in enumerate(it, start=2):
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue  # skip fully blank rows Excel often trails with
        fields = {header[i]: raw[i] if i < len(raw) else "" for i in range(len(header)) if header[i]}
        rows.append(_validate_row(line_number, fields, from_country))
    return rows


def _parse_csv(path: str, from_country: Optional[str] = None) -> list[BatchRow]:
    rows: list[BatchRow] = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        _check_header(reader.fieldnames)
        for line_number, raw_row in enumerate(reader, start=2):
            rows.append(_validate_row(line_number, raw_row, from_country))
    return rows


def parse_import(path: str, from_country: Optional[str] = None) -> list[BatchRow]:
    """Parse a recipient file, dispatching on extension (.xlsx or .csv).

    ``from_country`` is the sender's country, and is what makes a row
    international. Without it customs columns cannot be required, because
    whether they are needed at all depends on where the parcel is going *from*.
    """
    if Path(path).suffix.lower() in (".xlsx", ".xlsm"):
        return _parse_xlsx(path, from_country)
    return _parse_csv(path, from_country)


def revalidate(rows: list[BatchRow], from_country: Optional[str]) -> list[BatchRow]:
    """Re-run validation against a (possibly changed) sender country.

    The sender can be switched after a file is loaded, and that alone can turn
    every row international. Re-checking beats leaving a stale "5 valid, 0 with
    errors" on screen.
    """
    return [_validate_row(r.line_number, dict(r.fields), from_country) for r in rows]


# Back-compat alias: the import entry point was CSV-only before .xlsx support.
parse_csv = parse_import


def _row_customs_info(f: dict, from_country: Optional[str], declaration: Optional[dict]) -> Optional[dict]:
    """The declaration for one row, or None when the row is domestic.

    One customs item, built from the row's own columns; the parcel weight is
    reused as the item weight (see CSV_COLUMNS). The declaration-level fields —
    contents type, signer, non-delivery option — are the same for every row in
    the batch and come from the Batch page, not the spreadsheet.
    """
    if not is_international(from_country, f.get("to_country")):
        return None

    declaration = declaration or {}
    quantity = f.get("customs_quantity") or "1"
    item = customs_item(
        description=f.get("customs_description", ""),
        quantity=int(float(quantity)),
        value=float(f.get("customs_value") or 0),
        # The CSV's weight column is already ounces, which is what a customs
        # item wants, so it passes through rather than being converted.
        weight_oz=float(f["weight"]),
        origin_country=f.get("customs_origin_country") or (from_country or ""),
        currency=currency_for(from_country),
        hs_tariff_number=f.get("customs_hs_tariff", ""),
    )
    return build_customs_info([item], **declaration)


def _row_to_shipment_params(
    row: BatchRow,
    from_address_id: str,
    *,
    carrier: Optional[str] = None,
    service: Optional[str] = None,
    carrier_account_id: Optional[str] = None,
    delivery_confirmation: Optional[str] = None,
    insurance: Optional[str] = None,
    from_country: Optional[str] = None,
    declaration: Optional[dict] = None,
) -> dict:
    f = row.fields
    # A predefined package carries fixed dimensions, so send its code alone and
    # omit length/width/height — the same rule as a single Create Shipment
    # (app/services/shipments.py). Otherwise send the parcel's own dimensions.
    parcel = {"weight": float(f["weight"])}
    if f.get("predefined_package"):
        # The .xlsx dropdown offers carrier-qualified labels ("Royal Mail —
        # LETTER"); EasyPost's predefined_package wants the bare code alone. A
        # bare code typed straight into a CSV passes through unchanged.
        parcel["predefined_package"] = package_code_from_choice(f["predefined_package"])
    else:
        parcel.update({
            "length": float(f["length"]),
            "width": float(f["width"]),
            "height": float(f["height"]),
        })
    # Same printed-label format/size as a single shipment — label_size is only
    # honoured at creation time, so a batch has to carry it too. A signature
    # request rides along in the same options object.
    options = dict(preferred_label_options())
    if delivery_confirmation:
        options["delivery_confirmation"] = delivery_confirmation

    params = {
        "to_address": {
            "name": f.get("to_name") or None,
            "company": f.get("to_company") or None,
            "street1": f["to_street1"],
            "street2": f.get("to_street2") or None,
            "city": f["to_city"],
            # Optional: only some countries use one (STATE_REQUIRED_COUNTRIES).
            "state": f.get("to_state") or None,
            "zip": f["to_zip"],
            "country": f["to_country"],
            "phone": f.get("to_phone") or None,
            "email": f.get("to_email") or None,
        },
        "from_address": {"id": from_address_id},
        "parcel": parcel,
        "reference": f.get("reference") or None,
        "options": options,
    }

    # THE reason a batch can be bought at all. Unlike a single shipment — where
    # the user picks a rate from the rates table and buys that rate id — a batch
    # is never rated: `batch.buy` takes no body and EasyPost selects nothing for
    # you. The carrier and service must be declared here, at creation time, or
    # every shipment fails at purchase with "A carrier and service must be
    # provided to purchase through a Batch."
    if carrier:
        params["carrier"] = carrier
    if service:
        params["service"] = service
    # Documented alongside carrier/service. EasyPost accepted a batch without it
    # in testing, but a stale or wrong id is a hard error rather than a silent
    # skip, so it is sent only when known.
    if carrier_account_id:
        params["carrier_accounts"] = [carrier_account_id]
    if insurance:
        params["insurance"] = insurance

    # Only attachable at creation. A batch cannot be amended before purchase, so
    # a shipment that leaves here without a declaration can never acquire one —
    # it can only fail at the till.
    customs_info = _row_customs_info(f, from_country, declaration)
    if customs_info:
        params["customs_info"] = customs_info

    return params


def rate_representative_row(
    from_address_id: str,
    row: BatchRow,
    *,
    from_country: Optional[str] = None,
    declaration: Optional[dict] = None,
    delivery_confirmation: Optional[str] = None,
    insurance: Optional[str] = None,
):
    """Rate ONE row as an ordinary shipment, to find out what carriers quote.

    A batch is never rated, so the carrier and service have to be named before
    anything is priced — which means choosing them blind from a catalogue where
    Royal Mail alone publishes 243 services. Naming one the route does not
    support creates a batch that fails at purchase, and that failure is the
    expensive kind.

    A single shipment *is* rated, so this creates one from a real row and reads
    the answer. Deliberately built by the same `_row_to_shipment_params` the
    batch will use, minus the carrier and service: rating a differently-shaped
    payload would answer a question nobody asked. Creating a shipment costs
    nothing — only buying does — and an unbought shipment simply expires.

    Carrier and service are omitted on purpose. Naming them is what narrows
    EasyPost's reply to that one service instead of everything available.
    """
    params = _row_to_shipment_params(
        row,
        from_address_id,
        from_country=from_country,
        declaration=declaration,
        delivery_confirmation=delivery_confirmation,
        insurance=insurance,
    )
    client = client_manager.get_client()
    return client.shipment.create(**params)


def quoted_services(shipment) -> list[dict]:
    """Flatten a rated shipment into the carrier/service pairs it quoted.

    Only what the picker needs: which pairs exist, and what each costs so the
    list can show a price beside a name.
    """
    quotes = []
    for rate in getattr(shipment, "rates", None) or []:
        carrier = getattr(rate, "carrier", None)
        service = getattr(rate, "service", None)
        if not carrier or not service:
            continue
        quotes.append({
            "carrier": carrier,
            "service": service,
            "rate": getattr(rate, "rate", None),
            "currency": getattr(rate, "currency", None),
        })
    return quotes


def create_batch(
    from_address_id: str,
    rows: list[BatchRow],
    *,
    carrier: Optional[str] = None,
    service: Optional[str] = None,
    carrier_account_id: Optional[str] = None,
    delivery_confirmation: Optional[str] = None,
    insurance: Optional[str] = None,
    from_country: Optional[str] = None,
    declaration: Optional[dict] = None,
):
    """Create a batch. ``carrier`` and ``service`` are required to buy it later
    (see :func:`_row_to_shipment_params`); a batch created without them can be
    created but never purchased.

    ``from_country`` and ``declaration`` carry the customs details for a batch
    that crosses a border. Both are refused up front rather than half-built: a
    batch is not amendable, so an international shipment created without a
    declaration is dead on arrival and has already consumed a batch."""
    valid_rows = [r for r in rows if r.is_valid]
    if not valid_rows:
        raise ValueError("No valid rows to submit.")

    # Fail before the API call, not after. Creating the batch and discovering
    # this at purchase is the failure being fixed.
    if any(is_international(from_country, r.fields.get("to_country")) for r in valid_rows):
        if not declaration:
            raise ValueError("missing_customs_declaration")

    client = client_manager.get_client()
    shipments = [
        _row_to_shipment_params(
            r,
            from_address_id,
            carrier=carrier,
            service=service,
            carrier_account_id=carrier_account_id,
            delivery_confirmation=delivery_confirmation,
            insurance=insurance,
            from_country=from_country,
            declaration=declaration,
        )
        for r in valid_rows
    ]
    return client.batch.create(shipments=shipments)


def retrieve_batch(batch_id: str):
    client = client_manager.get_client()
    return client.batch.retrieve(batch_id)


def buy_batch(batch_id: str):
    client = client_manager.get_client()
    return client.batch.buy(batch_id)


def generate_batch_label(batch_id: str, file_format: str = "PDF"):
    client = client_manager.get_client()
    return client.batch.label(batch_id, file_format=file_format)


def _batch_state(batch) -> Optional[str]:
    return getattr(batch, "state", None) or getattr(batch, "status", None)


def batch_failure_messages(batch) -> list[str]:
    """Per-shipment failure messages carried inside a batch.

    Batch creation and purchase are asynchronous and never raise for a shipment
    that failed: the call returns normally and the batch reports the failure in
    each shipment's ``batch_status``/``batch_message``. A caller that only
    watches for exceptions therefore sees a batch that bought nothing and says
    nothing about why. This is where the explanation lives — for example
    "RoyalMailV3 does not offer service RoyalMail2ndClassSignedFor for this
    shipment", which is what a missing signature option produces.
    """
    messages: list[str] = []
    for shipment in getattr(batch, "shipments", None) or []:
        status = shipment.get("batch_status")
        if status not in ("creation_failed", "postage_purchase_failed"):
            continue
        message = shipment.get("batch_message") or status
        reference = shipment.get("reference") or shipment.get("id") or ""
        messages.append(f"{reference}: {message}" if reference else str(message))
    # One repeated cause across 200 shipments should read as one line, not 200.
    seen: set[str] = set()
    unique: list[str] = []
    for message in messages:
        body = message.split(": ", 1)[-1]
        if body in seen:
            continue
        seen.add(body)
        unique.append(message)
    return unique


def bought_shipment_ids(batch) -> list[str]:
    """Ids of the shipments in this batch that actually bought postage."""
    return [
        shipment.get("id")
        for shipment in getattr(batch, "shipments", None) or []
        if shipment.get("id") and shipment.get("tracking_code")
    ]


def full_shipments(batch) -> list:
    """Retrieve each bought shipment in full.

    The shipments embedded in a batch are **stubs**. Verified against a real
    purchased batch, each one carries only ``id``, ``reference``,
    ``tracking_code``, ``batch_status`` and ``batch_message`` — no
    ``postage_label``, no ``tracker``, no ``carrier``. Anything that needs a
    label URL or a tracker has to fetch the shipment itself, which is what this
    does. One request per shipment, so callers should treat it as a network
    operation and keep it off the UI thread.
    """
    client = client_manager.get_client()
    shipments = []
    for shipment_id in bought_shipment_ids(batch):
        try:
            shipments.append(client.shipment.retrieve(shipment_id))
        except Exception:
            logger.exception("Could not retrieve batch shipment %s", shipment_id)
    return shipments


def batch_label_urls(batch) -> list[str]:
    """Per-shipment label URLs for a bought batch, for the print sheet.

    Reads them off fully retrieved shipments rather than the batch's own stub
    entries, which carry no ``postage_label`` at all — checking the stubs
    returned nothing every time, leaving the print-sheet export permanently
    disabled after a batch purchase.
    """
    urls = []
    for shipment in full_shipments(batch):
        label = getattr(shipment, "postage_label", None)
        url = getattr(label, "label_url", None) if label else None
        if url:
            urls.append(url)
    return urls


def track_batch_shipments(batch) -> int:
    """Record every bought shipment in the local tracking table.

    Buying a label always creates a tracker on EasyPost's side — that is not
    optional and there is no way to opt out of carrier tracking. What is
    optional, and what this does, is copying those trackers into the app's own
    Tracking page so a batch does not vanish the moment it is bought. Returns
    the number recorded.

    The tracker is taken from the retrieved shipment rather than created afresh
    from the tracking code. Creating one is both redundant — buying the label
    already made it — and rejected outright in test mode, where EasyPost accepts
    only its own EZ-prefixed test tracking numbers.

    Best effort by design: the labels are already paid for, so a local
    bookkeeping failure must never be reported as a failed purchase.
    """
    from app.services.tracking import save_tracker_locally

    recorded = 0
    for shipment in full_shipments(batch):
        tracker = getattr(shipment, "tracker", None)
        if tracker is None:
            continue
        try:
            save_tracker_locally(tracker)
            recorded += 1
        except Exception:
            logger.exception("Could not record tracker for shipment %s", shipment.id)
    return recorded


def save_batch_locally(batch, source_csv: str = "") -> None:
    mode = client_manager.active_mode
    with db_cursor() as cur:
        cur.execute(
            """
            INSERT INTO batches (id, mode, status, num_shipments, source_csv)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(id) DO UPDATE SET
                status=excluded.status, num_shipments=excluded.num_shipments
            """,
            (
                batch.id,
                mode,
                _batch_state(batch),
                getattr(batch, "num_shipments", None),
                source_csv or None,
            ),
        )


@dataclass
class BatchRecord:
    id: str
    mode: str
    status: Optional[str]
    num_shipments: Optional[int]
    source_csv: Optional[str]


_BATCH_FIELDS = [f for f in BatchRecord.__dataclass_fields__]


def list_batches() -> list[BatchRecord]:
    mode = client_manager.active_mode
    with db_cursor() as cur:
        cur.execute(
            "SELECT * FROM batches WHERE mode = ? ORDER BY created_at DESC", (mode,)
        )
        rows = cur.fetchall()
    return [BatchRecord(**{k: row[k] for k in _BATCH_FIELDS}) for row in rows]
