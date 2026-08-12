from openpyxl import load_workbook

from app.services.batches import (
    CSV_COLUMNS,
    _row_to_shipment_params,
    parse_csv,
    parse_import,
    write_csv_template,
    write_xlsx_template,
)


def test_template_round_trips_as_valid(tmp_path):
    path = tmp_path / "template.csv"
    write_csv_template(str(path))

    rows = parse_csv(str(path))

    assert len(rows) == 1
    assert rows[0].is_valid
    assert rows[0].fields["to_city"] == "Boston"


def test_missing_required_column_raises(tmp_path):
    path = tmp_path / "bad.csv"
    path.write_text("to_street1,to_city\n123 Main St,Boston\n", encoding="utf-8")

    try:
        parse_csv(str(path))
        assert False, "expected ValueError for missing required columns"
    except ValueError as exc:
        assert "missing required columns" in str(exc).lower()


def test_non_numeric_dimension_is_flagged(tmp_path):
    path = tmp_path / "rows.csv"
    header = ",".join(CSV_COLUMNS)
    # length "ten" is non-numeric; trailing empty package + reference columns.
    row = "Jane,,123 Main St,,Boston,MA,02110,US,,,ten,6,4,16,,"
    path.write_text(f"{header}\n{row}\n", encoding="utf-8")

    rows = parse_csv(str(path))

    assert len(rows) == 1
    assert not rows[0].is_valid
    assert any("length" in e for e in rows[0].errors)


def test_predefined_package_makes_dimensions_optional(tmp_path):
    """A row that names a carrier package needs no length/width/height."""
    path = tmp_path / "predef.csv"
    header = ",".join(CSV_COLUMNS)
    # No dimensions at all, but predefined_package = FlatRateEnvelope.
    row = "Jane,,123 Main St,,Boston,MA,02110,US,,,,,,16,FlatRateEnvelope,ref-1"
    path.write_text(f"{header}\n{row}\n", encoding="utf-8")

    rows = parse_csv(str(path))

    assert len(rows) == 1
    assert rows[0].is_valid, rows[0].errors
    assert rows[0].fields["predefined_package"] == "FlatRateEnvelope"


def test_row_without_package_still_requires_dimensions(tmp_path):
    path = tmp_path / "nopredef.csv"
    header = ",".join(CSV_COLUMNS)
    row = "Jane,,123 Main St,,Boston,MA,02110,US,,,,,,16,,ref-1"  # no dims, no package
    path.write_text(f"{header}\n{row}\n", encoding="utf-8")

    rows = parse_csv(str(path))

    assert not rows[0].is_valid
    assert {"length", "width", "height"}.issubset(set(rows[0].errors))


def test_predefined_package_param_omits_dimensions():
    """The parcel sent to EasyPost carries the code alone, not dimensions."""
    from app.services.batches import BatchRow

    row = BatchRow(
        line_number=2,
        fields={
            "to_street1": "123 Main St", "to_city": "Boston", "to_state": "MA",
            "to_zip": "02110", "to_country": "US", "weight": "16",
            "predefined_package": "FlatRateEnvelope",
        },
        errors=[],
    )
    params = _row_to_shipment_params(row, "adr_123")
    assert params["parcel"] == {"weight": 16.0, "predefined_package": "FlatRateEnvelope"}
    assert "length" not in params["parcel"]


def _gb_row(**over):
    from app.services.batches import BatchRow

    fields = {
        "to_street1": "10 Downing St", "to_city": "London", "to_state": "",
        "to_zip": "SW1A 2AA", "to_country": "GB", "weight": "3.5",
        "predefined_package": "LETTER",
    }
    fields.update(over)
    return BatchRow(line_number=2, fields=fields, errors=[])


def test_gb_row_without_state_is_valid(tmp_path):
    """EasyPost's own documented GB example carries no `state`; requiring one
    rejected valid UK rows outright."""
    path = tmp_path / "gb.csv"
    header = ",".join(CSV_COLUMNS)
    row = "Jane,,10 Downing St,,London,,SW1A 2AA,GB,,,,,,3.5,LETTER,ref-gb"
    path.write_text(f"{header}\n{row}\n", encoding="utf-8")

    rows = parse_csv(str(path))

    assert len(rows) == 1
    assert rows[0].is_valid, rows[0].errors


def test_us_row_still_requires_state(tmp_path):
    path = tmp_path / "us.csv"
    header = ",".join(CSV_COLUMNS)
    row = "Jane,,123 Main St,,Boston,,02110,US,,,,,,16,,ref-us"
    path.write_text(f"{header}\n{row}\n", encoding="utf-8")

    rows = parse_csv(str(path))

    assert not rows[0].is_valid
    assert "to_state" in rows[0].errors


def test_carrier_and_service_are_top_level_not_in_options():
    """Batch buy fails without these: EasyPost never rates batch shipments and
    `batch.buy` takes no body, so the service must be declared at create time."""
    params = _row_to_shipment_params(
        _gb_row(), "adr_uk",
        carrier="RoyalMailV3", service="RoyalMail2ndClassSignedFor",
        carrier_account_id="ca_test", delivery_confirmation="SIGNATURE",
        insurance="20.00",
    )
    assert params["carrier"] == "RoyalMailV3"
    assert params["service"] == "RoyalMail2ndClassSignedFor"
    assert params["carrier_accounts"] == ["ca_test"]
    assert params["insurance"] == "20.00"
    # Signature belongs in options; carrier/service must NOT be nested there.
    assert params["options"]["delivery_confirmation"] == "SIGNATURE"
    assert "carrier" not in params["options"]
    assert "service" not in params["options"]


def test_absent_carrier_service_are_omitted_not_null():
    params = _row_to_shipment_params(_gb_row(), "adr_uk")
    for key in ("carrier", "service", "carrier_accounts", "insurance"):
        assert key not in params
    assert "delivery_confirmation" not in params["options"]


def test_blank_state_is_sent_as_none():
    params = _row_to_shipment_params(_gb_row(), "adr_uk")
    assert params["to_address"]["state"] is None


def test_carrier_qualified_package_reduces_to_bare_code():
    """A dropdown choice like "Royal Mail — LETTER" is sent to EasyPost as the
    bare code "LETTER", not the whole carrier-qualified label."""
    from app.services.batches import BatchRow

    row = BatchRow(
        line_number=2,
        fields={
            "to_street1": "10 Downing St", "to_city": "London", "to_state": "",
            "to_zip": "SW1A 2AA", "to_country": "GB", "weight": "3.5",
            "predefined_package": "Royal Mail — LETTER",
        },
        errors=[],
    )
    params = _row_to_shipment_params(row, "adr_uk")
    assert params["parcel"]["predefined_package"] == "LETTER"


def test_xlsx_template_has_package_dropdown(tmp_path):
    path = tmp_path / "template.xlsx"
    choices = ["FlatRateEnvelope", "Parcel", "SmallFlatRateBox"]
    write_xlsx_template(str(path), choices)

    # Round-trips through the importer as one valid row.
    rows = parse_import(str(path))
    assert len(rows) == 1
    assert rows[0].is_valid
    assert rows[0].fields["to_city"] == "Boston"

    # A data validation exists on the predefined_package column, sourced from a
    # hidden Packages sheet holding every choice.
    wb = load_workbook(str(path))
    ws = wb["Recipients"]
    assert len(ws.data_validations.dataValidation) == 1
    dv = ws.data_validations.dataValidation[0]
    assert "Packages" in dv.formula1
    opts = wb["Packages"]
    assert opts.sheet_state == "hidden"
    listed = [row[0].value for row in opts.iter_rows(min_row=2)]
    assert listed == choices


def test_xlsx_template_without_choices_has_no_dropdown(tmp_path):
    path = tmp_path / "template.xlsx"
    write_xlsx_template(str(path), [])

    wb = load_workbook(str(path))
    ws = wb["Recipients"]
    assert len(ws.data_validations.dataValidation) == 0
    assert "Packages" not in wb.sheetnames
    # Still a usable template.
    rows = parse_import(str(path))
    assert len(rows) == 1 and rows[0].is_valid
