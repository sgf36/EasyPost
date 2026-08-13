"""An international batch must carry a customs declaration.

The failure this pins: a GB sender importing five US recipients got "5 rows
loaded — 5 valid, 0 with errors", created the batch against a live account, and
only then saw every label fail with the carrier's wording —

    At least one item per package must be provided.

Nothing said "customs". The rows had passed validation, the batch existed, and
a batch cannot be amended, so the only way out was to start again. The cause was
structural: `customs_info` was assembled inside CreateShipmentView, so the batch
path had no access to it and no test asserted it should.
"""

import pytest

from app.services.batches import (
    CSV_COLUMNS,
    _row_to_shipment_params,
    _validate_row,
    create_batch,
    parse_import,
    revalidate,
    write_csv_template,
)

DECLARATION = {"customs_signer": "S Fields", "contents_type": "merchandise"}


def _row(**overrides):
    fields = {
        "to_name": "Karen Fields", "to_street1": "1 Sunset Blvd",
        "to_city": "Los Angeles", "to_state": "CA", "to_zip": "90001",
        "to_country": "US", "weight": "3.5", "predefined_package": "Letter",
    }
    fields.update(overrides)
    return fields


def _customs_row(**overrides):
    customs = {
        "customs_description": "Cotton t-shirt",
        "customs_quantity": "2",
        "customs_value": "12.50",
        "customs_hs_tariff": "610910",
    }
    customs.update(overrides)
    return _row(**customs)


def test_template_carries_the_customs_columns():
    for column in ("customs_description", "customs_value", "customs_origin_country"):
        assert column in CSV_COLUMNS


def test_template_still_round_trips(tmp_path):
    path = tmp_path / "t.csv"
    write_csv_template(str(path))
    rows = parse_import(str(path), from_country="GB")
    assert rows[0].is_valid, rows[0].errors


def test_domestic_row_needs_no_customs():
    row = _validate_row(2, _row(), from_country="US")
    assert row.is_valid, row.errors


def test_international_row_without_customs_is_rejected():
    row = _validate_row(2, _row(), from_country="GB")
    assert not row.is_valid
    assert "customs_description" in row.errors
    assert "customs_value" in row.errors


def test_international_row_with_customs_is_accepted():
    row = _validate_row(2, _customs_row(), from_country="GB")
    assert row.is_valid, row.errors


def test_unknown_sender_country_does_not_demand_customs():
    """A file can be parsed before a sender is chosen; guessing would demand a
    declaration for a domestic parcel."""
    row = _validate_row(2, _row(), from_country=None)
    assert row.is_valid, row.errors


def test_switching_sender_country_revalidates():
    rows = [_validate_row(2, _row(), from_country="US")]
    assert rows[0].is_valid
    again = revalidate(rows, from_country="GB")
    assert not again[0].is_valid
    assert "customs_description" in again[0].errors


def test_domestic_shipment_carries_no_customs_info():
    params = _row_to_shipment_params(
        _validate_row(2, _row(), from_country="US"), "adr_1",
        from_country="US", declaration=DECLARATION,
    )
    assert "customs_info" not in params


def test_international_shipment_carries_a_full_declaration():
    params = _row_to_shipment_params(
        _validate_row(2, _customs_row(), from_country="GB"), "adr_1",
        from_country="GB", declaration=DECLARATION,
    )
    info = params["customs_info"]
    assert info["customs_certify"] is True
    assert info["customs_signer"] == "S Fields"
    assert info["contents_type"] == "merchandise"
    assert info["non_delivery_option"] == "return"
    assert info["eel_pfc"]                      # some carriers reject a label without one

    item, = info["customs_items"]
    assert item["description"] == "Cotton t-shirt"
    assert item["quantity"] == 2
    assert item["value"] == 12.50
    assert item["hs_tariff_number"] == "610910"
    # The parcel weight doubles as the item weight, already in ounces.
    assert item["weight"] == 3.5


def test_declared_value_uses_the_senders_currency():
    """A London sender entering 12.50 means pounds. Declaring dollars misstates
    a customs document."""
    params = _row_to_shipment_params(
        _validate_row(2, _customs_row(), from_country="GB"), "adr_1",
        from_country="GB", declaration=DECLARATION,
    )
    assert params["customs_info"]["customs_items"][0]["currency"] == "GBP"


def test_origin_country_defaults_to_the_sender():
    params = _row_to_shipment_params(
        _validate_row(2, _customs_row(), from_country="GB"), "adr_1",
        from_country="GB", declaration=DECLARATION,
    )
    assert params["customs_info"]["customs_items"][0]["origin_country"] == "GB"


def test_blank_tariff_number_is_omitted_not_nulled():
    row = _validate_row(2, _customs_row(customs_hs_tariff=""), from_country="GB")
    params = _row_to_shipment_params(row, "adr_1", from_country="GB", declaration=DECLARATION)
    assert "hs_tariff_number" not in params["customs_info"]["customs_items"][0]


def test_creating_an_international_batch_without_a_declaration_is_refused():
    """Before the API call. A batch cannot be amended, so discovering this at
    purchase means the batch is already spent."""
    rows = [_validate_row(2, _customs_row(), from_country="GB")]
    with pytest.raises(ValueError, match="missing_customs_declaration"):
        create_batch("adr_1", rows, carrier="RoyalMailV3",
                     service="InternationalStandardOnAccount", from_country="GB")


def test_uk_is_not_a_country_code():
    """The one everybody types. EasyPost rejects it, and before reaching the
    API it also makes a London-to-London parcel look international, because
    "GB" != "UK" — so the row is asked for customs details it does not need."""
    row = _validate_row(2, _row(to_country="UK"), from_country="GB")
    assert not row.is_valid
    assert any("to_country is not a country code" in e for e in row.errors)


def test_a_country_name_is_not_a_code():
    row = _validate_row(2, _row(to_country="United Kingdom"), from_country="GB")
    assert not row.is_valid


def test_lowercase_and_dropdown_labels_both_normalise():
    """The spreadsheet dropdown offers "GB — United Kingdom"; a CSV gets typed
    by hand. Both reduce to the bare code before anything else reads it."""
    for value in ("gb", "GB — United Kingdom", " gb "):
        row = _validate_row(2, _row(to_country=value), from_country="GB")
        assert row.is_valid, (value, row.errors)
        assert row.fields["to_country"] == "GB"
        # And a domestic GB row stays domestic, so no customs is demanded.
        assert "customs_description" not in row.errors


def test_a_bad_origin_country_is_caught_too():
    row = _validate_row(2, _customs_row(customs_origin_country="UK"), from_country="GB")
    assert not row.is_valid
    assert any("customs_origin_country is not a country code" in e for e in row.errors)


def test_the_xlsx_template_offers_country_dropdowns(tmp_path):
    from openpyxl import load_workbook

    from app.services.batches import write_xlsx_template

    path = tmp_path / "t.xlsx"
    write_xlsx_template(str(path), package_choices=["Royal Mail — Letter"])
    wb = load_workbook(str(path))
    assert {"Countries", "OriginCountries"} <= set(wb.sheetnames)
    assert wb["Countries"].max_row - 1 == 197
    ranges = {str(dv.sqref) for dv in wb["Recipients"].data_validations.dataValidation}
    assert len(ranges) == 3, ranges       # packages, destination, origin


def test_the_reported_case(tmp_path):
    """Five US recipients from a London sender, no customs columns filled in —
    the exact import that reached purchase_failed in production."""
    header = ",".join(CSV_COLUMNS)
    blank = {c: "" for c in CSV_COLUMNS}
    lines = [header]
    for city in ("Tarzana", "Los Angeles", "Los Angeles", "Los Angeles", "Beverly Hills"):
        row = dict(blank, to_name="A Recipient", to_street1="1 Main St", to_city=city,
                   to_state="CA", to_zip="90001", to_country="US", weight="3.5",
                   predefined_package="Letter")
        lines.append(",".join(row[c] for c in CSV_COLUMNS))
    path = tmp_path / "recipients.csv"
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    rows = parse_import(str(path), from_country="GB")
    assert len(rows) == 5
    assert all(not r.is_valid for r in rows), "every international row should be flagged"
    assert all("customs_description" in r.errors for r in rows)
